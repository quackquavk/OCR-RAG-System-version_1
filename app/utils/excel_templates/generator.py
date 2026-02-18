from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List, Any, Optional, Union

class FormattedValue:
    """Helper class to encapsulate a value and its desired Excel number format."""
    def __init__(self, value: Any, number_format: str):
        self.value = value
        self.number_format = number_format

class ExcelGenerator:
    """Generate beautifully formatted Excel files from document data"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        
    def create_workbook(self) -> Workbook:
        """Create a new workbook"""
        self.wb = Workbook()
        self.ws = self.wb.active
        return self.wb

    def set_column_width(self, column: str, width: int) -> None:
        """Set column width"""
        self.ws.column_dimensions[column].width = width
    
    def merge_and_write(self, start_cell: str, end_cell: str, value: Any,
                       font: Optional[Font] = None,
                       alignment: Optional[Alignment] = None,
                       fill: Optional[PatternFill] = None) -> None:
        """Merge cells and write value with formatting"""
        self.ws.merge_cells(f'{start_cell}:{end_cell}')
        cell = self.ws[start_cell]
        
        if isinstance(value, FormattedValue):
            cell.value = value.value
            cell.number_format = value.number_format
        else:
            cell.value = value
            
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill
    
    def write_cell(self, cell: str, value: Any,
                  font: Optional[Font] = None,
                  alignment: Optional[Alignment] = None,
                  fill: Optional[PatternFill] = None,
                  border: Optional[Border] = None) -> None:
        """Write to a cell with formatting"""
        target_cell = self.ws[cell]
        
        if isinstance(value, FormattedValue):
            target_cell.value = value.value
            target_cell.number_format = value.number_format
        else:
            target_cell.value = value
            
        if font:
            target_cell.font = font
        if alignment:
            target_cell.alignment = alignment
        if fill:
            target_cell.fill = fill
        if border:
            target_cell.border = border
    
    def apply_border_to_range(self, start_cell: str, end_cell: str, border: Border) -> None:
        """Apply border to a range of cells"""
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(f'{start_cell}:{end_cell}')
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = border
    
    def get_bytes(self) -> bytes:
        """Get Excel file as bytes"""
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def create_font(bold: bool = False, size: int = 11, color: str = "000000", name: str = "Calibri"):
        """Create a font style"""
        return Font(name=name, size=size, bold=bold, color=color)
    
    @staticmethod
    def create_alignment(horizontal: str = "left", vertical: str = "center", wrap_text: bool = False):
        """Create alignment style"""
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
    
    @staticmethod
    def create_fill(color: str = "FFFFFF"):
        """Create fill style"""
        return PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    @staticmethod
    def create_border(style: str = "thin", color: str = "000000"):
        """Create border style"""
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    
    def format_currency(self, value: Any) -> Union[FormattedValue, str]:
        """Format value as currency"""
        try:
            if value is None or value == "":
                return FormattedValue(0.0, '#,##0.00')
                
            if isinstance(value, str):
                cleaned_value = value.replace("$", "").replace(",", "").strip()
                if not cleaned_value:
                     return FormattedValue(0.0, '#,##0.00')
                value = float(cleaned_value)
                
            if isinstance(value, (int, float)):
                 return FormattedValue(float(value), '#,##0.00')
                 
            return str(value)
        except (ValueError, TypeError):
            return str(value)
    
    def format_date(self, value: Any) -> str:
        """Format date value"""
        if not value:
            return ""
        
        if isinstance(value, str):
            return value
        
        if isinstance(value, datetime):
            return value.strftime("%B %d, %Y")
        
        return str(value)
    
    def auto_adjust_column_width(self, column: str, min_width: int = 10, max_width: int = 50) -> None:
        """Auto-adjust column width based on content"""
        max_length = 0
        column_letter = column
        
        for cell in self.ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        self.ws.column_dimensions[column_letter].width = adjusted_width

